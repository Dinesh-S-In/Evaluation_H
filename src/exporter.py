"""Export ranked evaluation results to CSV and formatted Excel."""

from __future__ import annotations

import logging
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.models import RankedEvaluation

logger = logging.getLogger(__name__)

_MAX_EXCEL_COL_WIDTH: int = 55


def build_results_dataframe(ranked: list[RankedEvaluation]) -> pd.DataFrame:
    """Flatten ranked evaluations into a single tabular DataFrame."""
    return pd.DataFrame.from_records(records_from_ranked(ranked))


def records_from_ranked(ranked: list[RankedEvaluation]) -> list[dict[str, Any]]:
    """Convert ranked evaluations into JSON-serializable row dicts."""
    rows: list[dict[str, Any]] = []
    for r in ranked:
        sub = r.submission
        ev = r.evaluation
        b = ev.breakdown
        final_total = r.manual_total_score if r.manual_total_score is not None else ev.total_score
        final_sl = r.effective_shortlisted()
        rows.append(
            {
                "submission_id": sub.submission_id,
                "team_name": sub.team_name or "",
                "row_index": sub.row_index,
                "rank": r.rank,
                "shortlisted": bool(r.shortlisted),
                "manual_shortlisted": r.manual_shortlisted,
                "final_shortlisted": bool(final_sl),
                "similarity_flag": bool(r.similarity_flag),
                "similar_submission_ids": ";".join(r.similar_submission_ids),
                "category_valid": bool(sub.category_valid),
                "category_warning": sub.category_warning or "",
                "Key Submission Impact": sub.key_submission_impact,
                "Business Use Case and Impact": sub.business_use_case_and_impact,
                "Solution/Approach Overview": sub.solution_approach_overview,
                "Category Selection": sub.category_selection,
                "Tools and Technology Used": sub.tools_and_technology_used,
                "business_impact_score": b.business_impact.score,
                "business_impact_reason": b.business_impact.reason,
                "feasibility_scalability_score": b.feasibility_scalability.score,
                "feasibility_scalability_reason": b.feasibility_scalability.reason,
                "ai_depth_creativity_score": b.ai_depth_creativity.score,
                "ai_depth_creativity_reason": b.ai_depth_creativity.reason,
                "total_score": ev.total_score,
                "manual_total_score": r.manual_total_score,
                "final_total_score": final_total,
                "model_reported_total": ev.model_reported_total,
                "final_summary": ev.final_summary,
                "shortlist_recommendation": ev.shortlist_recommendation,
                "evaluator_notes": r.evaluator_notes or "",
                "manual_edits": r.has_manual_edits(),
            }
        )
    return rows


def _write_formatted_worksheet(df: pd.DataFrame, writer: pd.ExcelWriter) -> None:
    """Write a single formatted worksheet named ``results``."""
    df.to_excel(writer, index=False, sheet_name="results")
    ws = writer.sheets["results"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def _autosize_columns(ws: Any) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:  # noqa: BLE001
                val = ""
            max_length = max(max_length, min(len(val), 80))
        width = min(_MAX_EXCEL_COL_WIDTH, max(12, max_length + 2))
        ws.column_dimensions[column_letter].width = width


def export_dataframe_to_excel(df: pd.DataFrame, path: Path) -> None:
    """Write a DataFrame to xlsx with bold header row and reasonable column widths."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_formatted_worksheet(df, writer)
    logger.info("Wrote Excel export to %s", path)


def export_dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Serialize a formatted workbook to bytes (for browser downloads)."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_formatted_worksheet(df, writer)
    return buffer.getvalue()


def export_dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Serialize a DataFrame to UTF-8 CSV bytes with a BOM for Excel compatibility."""
    buf = BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue()


def export_ranked_results(
    ranked: list[RankedEvaluation],
    *,
    full_csv: Path,
    full_xlsx: Path,
    top10_csv: Path,
    top10_xlsx: Path,
) -> None:
    """
    Export full results and top-ten shortlist to CSV and Excel.

    Args:
        ranked: Ordered ranked evaluations (best rank first).
        full_csv / full_xlsx: destinations for all rows.
        top10_csv / top10_xlsx: destinations for shortlisted rows only.
    """
    df = build_results_dataframe(ranked)
    full_csv.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(full_csv, index=False)
    export_dataframe_to_excel(df, full_xlsx)

    if "shortlisted" in df.columns:
        mask = df["shortlisted"].fillna(False).astype(bool)
        top = df.loc[mask].copy()
    else:
        top = df.head(10).copy()

    top.to_csv(top10_csv, index=False)
    export_dataframe_to_excel(top, top10_xlsx)

    logger.info(
        "Exported full_results (%s rows) and top10 (%s rows)",
        len(df),
        len(top),
    )
